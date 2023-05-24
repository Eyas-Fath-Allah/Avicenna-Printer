import os
import PyPDF2
from openpyxl import workbook, load_workbook


def get_pdf_files(path, i):
    files_path_with_new_file_name = {}
    for x in range(3):
        name = ('0' * x) + str(i)
        for f in os.listdir(path):
            if os.path.splitext(f)[1] == '.pdf' and name in os.path.splitext(f)[0]:
                new_file_name = os.path.splitext(f)[0]
                file_path = os.path.join(path, f)
                files_path_with_new_file_name[file_path] = new_file_name
    return files_path_with_new_file_name


def extract_all_pdf_text(pdf_path):
    with open(pdf_path, 'rb') as pdf_file:
        reader = PyPDF2.PdfReader(pdf_file)
        all_text = []
        for i in range(len(reader.pages)):
            all_text.append(reader.pages[i].extract_text())
        text = " ".join(all_text)
        return text.splitlines()


def correct_the_errors(pdf_text, error_list):
    for error in error_list:
        for index, line in enumerate(pdf_text):
            if error in line:
                corrected_line = line.replace(error, '')
                pdf_text[index] = corrected_line


def extract_search_dictionary_from_text(pdf_text, errors, search_dictionary):
    result = {}
    correct_the_errors(pdf_text, errors)
    for key in search_dictionary.keys():
        for index1, line in enumerate(pdf_text):
            if key == 'KREATİNİN-SPOT İDRAR':
                q = get_KREATİNİN_SPOT_İDRAR(pdf_text)
                if q:
                    result[key] = q
                continue
            if key == 'Adı Soyadı':
                result[key] = get_name(pdf_text)
                continue
            if 'Adı' in line:
                continue
            if (key == 'KREATİNİN' or key == 'KALSİYUM' or key == 'ALBÜMİN' or key == 'FOSFOR') and 'İDRAR' in line:
                continue
            if key == 'ÜRE' and 'TOPLAMA' in line:
                continue
            if key in line:
                index = line.index(key)
                result[key] = line[: index - 1].replace(" ", "").replace(",", ".")
                del pdf_text[index1]
                if "*" in result[key]:
                    result[key] = result[key].replace("*", "")
                    if key in convert_1000:
                        result[key] = str(int(float(result[key]) * 1000))
                    if key == 'İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN':
                        result[key] = str(float(result[key]) / 1000)
                else:
                    if key in convert_1000:
                        result[key] = str(int(float(result[key]) * 1000))
                    if key == 'İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN':
                        if '<' in result[key] or '>' in result[key]:
                            first_letter = result[key][0]
                            result[key] = result[key][1:]
                            result[key] = first_letter + str(float(result[key]) / 1000)
                        else:
                            result[key] = str(float(result[key]) / 1000)
                break
    if 'İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN' not in result.keys():
        if 'KREATİNİN-SPOT İDRAR' in result.keys() and 'PROTEİN KANTİTATİF (İDRARDA)SPOT' in result.keys():
            get_İDRARDA_PROTEİN_SPOT_division_İDRARDA_KREATİNİN(result)
    return result


def get_KREATİNİN_SPOT_İDRAR(pdf_text):
    x = ''
    for line in pdf_text:
        if 'KREATİNİN-SPOT İDRAR' in line or 'KREATİNİN SPOT İDRAR' in line:
            if '*' not in line:
                index = line.index('KREATİNİN')
                x = line[: index - 1].replace(" ", "").replace(",", ".")
    if x != '':
        return x


def get_İDRARDA_PROTEİN_SPOT_division_İDRARDA_KREATİNİN(dic):
    if 'KREATİNİN-SPOT İDRAR' in dic.keys() and 'PROTEİN KANTİTATİF (İDRARDA)SPOT' in dic.keys():
        if '<' in dic['KREATİNİN-SPOT İDRAR'] or '>' in dic['KREATİNİN-SPOT İDRAR']:
            first_letter1 = dic['KREATİNİN-SPOT İDRAR'][0]
            dic['KREATİNİN-SPOT İDRAR'] = dic['KREATİNİN-SPOT İDRAR'][1:]
            if '<' in dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'] or '>' in dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']:
                first_letter2 = dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'][0]
                dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'] = dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'][1:]
                divide = float(dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']) / float(dic['KREATİNİN-SPOT İDRAR'])
                dic['KREATİNİN-SPOT İDRAR'] = first_letter1 + dic['KREATİNİN-SPOT İDRAR']
                dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'] = first_letter2 + dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']
                dic['İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN'] = str(divide)[:6]
            else:
                divide = float(dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']) / float(dic['KREATİNİN-SPOT İDRAR'])
                dic['KREATİNİN-SPOT İDRAR'] = first_letter1 + dic['KREATİNİN-SPOT İDRAR']
                dic['İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN'] = str(divide)[:6]
        else:
            if '<' in dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'] or '>' in dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']:
                first_letter2 = dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'][0]
                dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'] = dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'][1:]
                divide = float(dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']) / float(dic['KREATİNİN-SPOT İDRAR'])
                dic['PROTEİN KANTİTATİF (İDRARDA)SPOT'] = first_letter2 + dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']
                dic['İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN'] = str(divide)[:6]
            else:
                divide = float(dic['PROTEİN KANTİTATİF (İDRARDA)SPOT']) / float(dic['KREATİNİN-SPOT İDRAR'])
                dic['İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN'] = str(divide)[:6]

def get_name(pdf_text):
    for line in pdf_text:
        if 'Adı Soyadı' in line:
            name = line.replace('Adı Soyadı ', '')
            date = pdf_text[1][:10]
            name_date = name + '            ' + date
            return name_date


# Control Panel:
pdfs_path = 'C:\\Users\\user\\Desktop\\Canacenna Printer'
load_excel_path = 'C:\\Users\\user\\Desktop\\Canacenna Printer\\Canacenna\\YENİ KAN TABLOSU.xlsx'
result_excel_path = 'C:\\Users\\user\\Desktop\\Canacenna Printer\\Result.xlsx'


search_dictionary = {
    'Adı Soyadı': ['A1', 'H1', 'A27', 'H27'],
    'İDRARDA PROTEİN/24 SAAT': ['D23', 'K23', 'D49', 'K49'], 'İDRARDA KREATİNİN/24 SAAT': ['D24', 'K24', 'D50', 'K50'],
    'KREATİNİN-SPOT İDRAR': ['D21', 'K21', 'D47', 'K47'], 'PROTEİN KANTİTATİF (İDRARDA)SPOT': ['D20', 'K20', 'D46', 'K46'],
    'İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN': ['D22', 'K22', 'D48', 'K48'],
    'AÇLIK KAN GLUKOZU': ['B2', 'I2', 'B28', 'I28'], 'KREATİNİN': ['B3', 'I3', 'B29', 'I29'],
    'ÜRE': ['B4', 'I4', 'B30', 'I30'], 'BUN': ['B5', 'I5', 'B31', 'I31'],
    'ÜRİK ASİT': ['B6', 'I6', 'B32', 'I32'], 'SODYUM': ['B7', 'I7', 'B33', 'I33'],
    'POTASYUM': ['B8', 'I8', 'B34', 'I34'], 'KLOR': ['B9', 'I9', 'B35', 'I35'], 'KALSİYUM': ['B10', 'I10', 'B36', 'I36'],
    'FOSFOR': ['B11', 'I11', 'B37', 'I37'], 'MAGNEZYUM': ['B12', 'I12', 'B38', 'I38'],
    'TOTAL PROTEİN': ['B13', 'I13', 'B39', 'I39'], 'ALBÜMİN': ['B14', 'I14', 'B40', 'I40'],
    'DİREKT BİLİRUBİN': ['F13', 'M13', 'F39', 'M39'], 'TOTAL BİLİRUBİN': ['F15', 'M15', 'F41', 'M41'],
    'İNDİREKT BİLİRUBİN': ['F14', 'M14', 'F40', 'M40'], 'ASPARTAT AMİNOTRANSFERAZ': ['B15', 'I15', 'B41', 'I41'],
    'ALANİN AMİNOTRANSFERAZ': ['B16', 'I16', 'B42', 'I42'], 'ALKALEN FOSFATAZ': ['B17', 'I17', 'B43', 'I43'],
    'GAMMA GLUTAMİL TRANSFERAZ': ['B18', 'I18', 'B44', 'I44'], 'LAKTAT DEHİDROGENAZ': ['B19', 'I19', 'B45', 'I45'],
    'CRP': ['D11', 'K11', 'D37', 'K37'], 'SEDİMANTASYON': ['D12', 'K12', 'D38', 'K38'],
    'LÖKOSİT (WBC)': ['D2', 'K2', 'D28', 'K28'], 'TROMBOSİT': ['D8', 'K8', 'D34', 'K34'],
    'HEMOGLOBİN': ['D3', 'K3', 'D29', 'K29'], 'HEMATOKRIT': ['D4', 'K4', 'D30', 'K30'],
    'ORTALAMA ERITROSIT HACMİ (MCV)': ['D5', 'K5', 'D31', 'K31'],
    'ORTALAMA ERITROSIT HEMOGLOBİN': ['D6', 'K6', 'D32', 'K32'], 'ERİTROSİT DAĞILIM GENİŞLİĞİ (RDW)': ['D7', 'K7', 'D33', 'K33'],
    'NÖTROFIL SAYISI': ['D9', 'K9', 'D35', 'K35'], 'LENFOSIT SAYISI': ['D10', 'K10', 'D36', 'K36'],
    'FERRİTİN': ['D13', 'K13', 'D39', 'K39'], 'TRİGLİSERİD': ['B20', 'I20', 'B46', 'I46'],
    'KOLESTEROL VLDL': ['B21', 'I21', 'B47', 'I47'], 'HDL KOLESTEROL': ['B22', 'I22', 'B48', 'I48'],
    'TOTAL KOLESTEROL': ['B24', 'I24', 'B50', 'I50'], 'LDL KOLESTEROL': ['B23', 'I23', 'B49', 'I49'],
    'Transferrin Saturasyonu': ['D16', 'K16', 'D42', 'K42'],
    'TOTAL DEMİR BAĞLAMA KAPASİTESİ': ['D15', 'K15', 'D41', 'K41'], 'SERUM DEMİRİ': ['D14', 'K14', 'D40', 'K40'],
    'B12 VİTAMİNİ': ['D17', 'K17', 'D43', 'K43'], 'FOLİK ASİT': ['D18', 'K18', 'D44', 'K44'],
    'DANSİTE': ['F2', 'M2', 'F28', 'M28'], 'pH': ['F3', 'M3', 'F29', 'M29'], 'PROTEİN 0 mg/dL': ['F4', 'M4', 'F30', 'M30'],
    'GLUKOZ': ['F5', 'M5', 'F31', 'M31'], 'KETON': ['F8', 'M8', 'F34', 'M34'], 'BİLİRUBİN': ['F9', 'M9', 'F35', 'M35'],
    'KAN 0 e/µL': ['F6', 'M6', 'F32', 'M32'],
    'NİTRİT': ['F7', 'M7', 'F33', 'M33'], 'ÜROBİLİNOJEN': ['F11', 'M11', 'F37', 'M37'], 'LÖKOSİT ESTERAZ': ['F10', 'M10', 'F36', 'M36']
}

error_list = ['KAN GAZLARI+PH+NA+K+CA', 'SERUM DEMİRİ VE TOTAL DEMİR', '%SATURASYON (', 'KREATİNİN ÇOCUK (ADOLESAN',
              'GLİKOZİLE HEMOGLOBİN', 'PROTEİN/KREATİNİN', 'BİLİRUBİN (TOTAL+DİREKT)']

convert_1000 = ['LÖKOSİT (WBC)', 'TROMBOSİT', 'NÖTROFIL SAYISI', 'LENFOSIT SAYISI']

wb = load_workbook(load_excel_path)
ws = wb.active

for i in range(1, 5):
    files_path_with_new_file_name = get_pdf_files(pdfs_path, i)
    if files_path_with_new_file_name:
        for pdf_path, new_file_name in files_path_with_new_file_name.items():
            pdf_text = extract_all_pdf_text(pdf_path)
            # with open(f"C:\\Users\\user\\Desktop\\Canacenna Printer\\Test1.txt", "w", encoding='utf-8') as file:
            #     for line in pdf_text:
            #         file.write(f'{line}\n')
            result = extract_search_dictionary_from_text(pdf_text, error_list, search_dictionary)
            for key, value in result.items():
                cell = search_dictionary[key][i-1]
                ws[cell] = value


wb.save(result_excel_path)