import os
import PyPDF2
from openpyxl import workbook, load_workbook

def get_pdf_files(path, i):
    files_path_with_new_file_name = {}
    for x in range(3):
        name = ('0' * x) + str(i)
        for f in os.listdir(path):
            if os.path.splitext(f)[1] == '.pdf' and name in os.path.splitext(f)[0]:
                new_file_name = os.path.splitext(f)[0]
                file_path = os.path.join(path, f)
                files_path_with_new_file_name[file_path] = new_file_name
    return files_path_with_new_file_name


def extract_all_pdf_text(pdf_path):
    with open(pdf_path, 'rb') as pdf_file:
        reader = PyPDF2.PdfReader(pdf_file)
        all_text = []
        for i in range(len(reader.pages)):
            all_text.append(reader.pages[i].extract_text())
        text = " ".join(all_text)
        return text.splitlines()


def correct_the_errors(pdf_text, error_list):
    for error in error_list:
        for index, line in enumerate(pdf_text):
            if error in line:
                corrected_line = line.replace(error, '')
                pdf_text[index] = corrected_line


def extract_search_dictionary_from_text(pdf_text, errors, search_dictionary):
    result = {}
    correct_the_errors(pdf_text, errors)
    for key in search_dictionary.keys():
        for index1, line in enumerate(pdf_text):
            if key in line:
                index = line.index(key)
                result[key] = line[: index - 1].replace(" ", "").replace(",", ".")
                del pdf_text[index1]
                if "*" in result[key]:
                    result[key] = result[key].replace("*", "")
                    if prefix_with_asterisk:
                        result[key] = "* " + result[key]
                break
    return result


# Control Panel:
folder_path = 'C:\\Users\\eyas4\\Desktop\\Pediatri-Nefroloji\\Test'
version = 'v.3.0.0'
prefix_with_asterisk = False  # Default is True.



search_dictionary = {
    'KREATİNİN-SPOT İDRAR': ['D20', 'K20', 'D46', 'K46'], 'PROTEİN KANTİTATİF (İDRARDA)SPOT': ['D19', 'K19', 'D45', 'K45'],
    'İDRARDA PROTEİN SPOT / İDRARDA KREATİNİN': ['D21', 'K21', 'D47', 'K47'],
    'AÇLIK KAN GLUKOZU': ['B1', 'I1', 'B27', 'I27'], 'KREATİNİN': ['B2', 'I2', 'B28', 'I28'],
    'ÜRE': ['B3', 'I3', 'B29', 'I29'], 'BUN': ['B4', 'I4', 'B30', 'I30'],
    'HGFH': ['D24', 'K24', 'D50', 'K50'], 'ÜRİK ASİT': ['B5', 'I5', 'B31', 'I31'], 'SODYUM': ['B6', 'I6', 'B32', 'I32'],
    'POTASYUM': ['B7', 'I7', 'B33', 'I33'], 'KLOR': ['B8', 'I8', 'B34', 'I34'], 'KALSİYUM': ['B9', 'I9', 'B35', 'I35'],
    'FOSFOR': ['B10', 'I10', 'B36', 'I36'], 'MAGNEZYUM': ['B11', 'I11', 'B37', 'I37'],
    'TOTAL PROTEİN': ['B12', 'I12', 'B38', 'I38'], 'ALBÜMİN': ['B13', 'I13', 'B39', 'I39'],
    'DİREKT BİLİRUBİN': ['F12', 'M12', 'F38', 'M38'], 'TOTAL BİLİRUBİN': ['F14', 'M14', 'F40', 'M40'],
    'İNDİREKT BİLİRUBİN': ['F13', 'M13', 'F39', 'M39'], 'ASPARTAT AMİNOTRANSFERAZ': ['B14', 'I14', 'B40', 'I40'],
    'ALANİN AMİNOTRANSFERAZ': ['B15', 'I15', 'B41', 'I41'], 'ALKALEN FOSFATAZ': ['B16', 'I16', 'B42', 'I42'],
    'GAMMA GLUTAMİL TRANSFERAZ': ['B17', 'I17', 'B43', 'I43'], 'LAKTAT DEHİDROGENAZ': ['B18', 'I18', 'B44', 'I44'],
    'CRP': ['D10', 'K10', 'D36', 'K36'], 'SEDİMANTASYON': ['D11', 'K11', 'D37', 'K37'],
    'LÖKOSİT': ['D1', 'K1', 'D27', 'K27'], 'TROMBOSİT': ['D7', 'K7', 'D33', 'K33'],
    'HEMOGLOBİN': ['D2', 'K2', 'D28', 'K28'], 'HEMATOKRIT': ['D3', 'K3', 'D29', 'K29'],
    'ORTALAMA ERITROSIT HACMİ (MCV)': ['D4', 'K4', 'D30', 'K30'],
    'ORTALAMA ERITROSIT HEMOGLOBİN': ['D5', 'K5', 'D31', 'K31'], 'ERİTROSİT DAĞILIM GENİŞLİĞİ (RDW)': ['D6', 'K6', 'D32', 'K32'],
    'NÖTROFIL SAYISI': ['D8', 'K8', 'D34', 'K34'], 'LENFOSIT SAYISI': ['D9', 'K9', 'D35', 'K35'],
    'FERRİTİN': ['D12', 'K12', 'D38', 'K38'], 'TRİGLİSERİD': ['B19', 'I19', 'B45', 'I45'],
    'KOLESTEROL VLDL': ['B20', 'I20', 'B46', 'I46'], 'HDL KOLESTEROL': ['B21', 'I21', 'B47', 'I47'],
    'TOTAL KOLESTEROL': ['B23', 'I23', 'B49', 'I49'], 'LDL KOLESTEROL': ['B22', 'I22', 'B48', 'I48'],
    'Transferrin Saturasyonu': ['D15', 'K15', 'D41', 'K41'],
    'TOTAL DEMİR BAĞLAMA KAPASİTESİ': ['D14', 'K14', 'D40', 'K40'], 'SERUM DEMİRİ': ['D13', 'K13', 'D39', 'K39'],
    'B12 VİTAMİNİ': ['D16', 'K16', 'D42', 'K42'], 'FOLİK ASİT': ['D17', 'K17', 'D43', 'K43'],
    'DANSİTE': ['F1', 'M1', 'F27', 'M27'], 'pH': ['F2', 'M2', 'F28', 'M28'], 'PROTEİN': ['F3', 'M3', 'F29', 'M29'],
    'GLUKOZ': ['F4', 'M4', 'F30', 'M30'], 'KETON': ['F7', 'M7', 'F33', 'M33'], 'BİLİRUBİN': ['F8', 'M8', 'F34', 'M34'],
    'KAN': ['F5', 'M5', 'F31', 'M31'],
    'NİTRİT': ['F6', 'M6', 'F32', 'M32'], 'ÜROBİLİNOJEN': ['F10', 'M10', 'F36', 'M36'], 'LÖKOSİT ESTERAZ': ['F9', 'M9', 'F36', 'M36']
}

blood_gas = {'PH': '', 'PC02': '', 'PO2': '', 'SO2': ''}

synonyms_dictionary = {
    'SODYUM': 'NA', 'POTASYUM': 'K', 'KALSİYUM': 'Ca', 'KLOR': 'Cl', 'FOSFOR': 'P', 'MAGNEZYUM': 'Mg',
    'ASPARTAT AMİNOTRANSFERAZ': 'AST', 'ALANİN AMİNOTRANSFERAZ': 'ALT', 'ALKALEN FOSFATAZ': 'ALP',
    'GAMMA GLUTAMİL TRANSFERAZ': 'GGT', 'LAKTAT DEHİDROGENAZ': 'LDH', 'LÖKOSİT': 'WBC', 'ERİTROSİT': 'RBC',
    'TROMBOSİT': 'PLT', 'HEMOGLOBİN': 'HB', 'HEMATOKRIT': 'Htc',
    'NÖTROFIL SAYISI': 'Neu', 'LENFOSIT SAYISI': 'Lenfosit', 'MONOSIT SAYISI': 'Mono', 'EOZINOFIL SAYISI': 'Eozinofil',
    'BAZOFIL SAYISI': 'Bazo', 'TOTAL DEMİR BAĞLAMA KAPASİTESİ': 'TDBK',
    'B12 VİTAMİNİ': 'B12', 'SERBEST T4': 'sT4', 'SERBEST T3': 'sT3', 'KOLESTEROL VLDL': 'VLD', 'LDL KOLESTEROL': 'LDL',
    'NON-HDL KOLESTEROL': 'NON-HDL'
}

error_list = ['KAN GAZLARI+PH+NA+K+CA', 'SERUM DEMİRİ VE TOTAL DEMİR', '%SATURASYON (', 'KREATİNİN ÇOCUK (ADOLESAN',
              'GLİKOZİLE HEMOGLOBİN', 'PROTEİN/KREATİNİN', 'Örnek Türü :KAN']


wb = load_workbook('C:\\Users\\eyas4\\Desktop\\Pediatri-Nefroloji\\KAN TABLOSU.xlsx')
ws = wb.active

for i in range(1, 5):
    files_path_with_new_file_name = get_pdf_files(folder_path, i)
    if files_path_with_new_file_name:
        for pdf_path, new_file_name in files_path_with_new_file_name.items():
            pdf_text = extract_all_pdf_text(pdf_path)
            result = extract_search_dictionary_from_text(pdf_text, error_list, search_dictionary)
            for key, value in result.items():
                cell = search_dictionary[key][i-1]
                ws[cell] = value


wb.save('C:\\Users\\eyas4\\Desktop\\Pediatri-Nefroloji\\Test.xlsx')


